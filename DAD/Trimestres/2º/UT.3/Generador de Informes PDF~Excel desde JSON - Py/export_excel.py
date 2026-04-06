# Exportación de los datos a Excel utilizando "openpyxl".
    # Módulo para exportar datos de la API DummyJSON a un archivo Excel.
import os                                                               # Importa la biblioteca "os" para operaciones del sistema operativo.
from datetime import datetime                                           # Importa la clase "datetime" para manejar fechas y horas.

from openpyxl import Workbook                                           # Importa la clase Workbook para crear archivos Excel.
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Importa estilos para formato.
from openpyxl.utils import get_column_letter                            # Importa función para obtener letra de columna.



# Función/Método para exportar los datos a un archivo Excel.
    # La función recibe como parámetros la lista de productos y el término de búsqueda (si lo hay) y genera un archivo Excel con los datos.
def export_data_to_excel(productos, termino = ""):
    # Crea una carpeta/Directorio "Informes Excel generados" en el directorio actual.
    ruta_carpeta = os.path.join(os.getcwd(), "Informes Excel generados")    # Define la ruta de la carpeta.

    # Compreuba/Verifica previo a crear la carpeta si esta existe. Si no existe crea la carpeta.
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)                       # Crea la carpeta si no existe.
        print(f"\n\n\tCarpeta creada: {ruta_carpeta}")  # Mensaje de confirmación de creación de la carpeta por consola.
    # Si la carpeta ya existe, muestra un mensaje indicando/informando al usuario de su existencia.
    else:
        print(f"\n\n\tLa carpeta ya existe: {ruta_carpeta}")  # Mensaje de confirmación de que la carpeta ya existe por consola.

    # Obtiene la fecha y hora actual para usarla en el nombre del archivo.
    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Compreuba/Verifica si hay un término/parámetro de búsqueda para definir el nombre del archivo y el título del informe.
    if termino:
        nombre_archivo = f"Informe_productos_DummyJSON_{termino}_{fecha_actual}.xlsx"   # Nombre del archivo con término/parámetro de búsqueda.
        titulo = f"Informe de productos - DummyJSON - Búsqueda: '{termino}'"            # Título del informe con término/parámetro de búsqueda.
    # Si no hay término/parámetro de búsqueda, indicamos que es un informe de todos los productos.
    else:
        nombre_archivo = f"Informe_todos_los_productos_DummyJSON_{fecha_actual}.xlsx"   # Nombre del archivo para todos los productos.
        titulo = "Informe de productos - DummyJSON - Listado completo"                  # Título del informe para todos los productos.

    # Ruta completa del archivo Excel.
    ruta_archivo = os.path.join(ruta_carpeta, nombre_archivo)

    # Crea el libro de trabajo (workbook) de Excel.
    wb = Workbook()                     # Crea un nuevo libro de trabajo.
    ws = wb.active                      # Selecciona y/u obtiene la hoja activa.
    ws.title = "Productos DummyJSON"    # Establece y/u asigna el título de la hoja.

    # Define los estilos para el Excel.
    title_font = Font(name = 'Arial', size = 16, bold = True, color = "2C3E50") # Estilo/Formato del título principal.
    title_alignment = Alignment(horizontal = "center", vertical = "center")     # Alineación del título principal.
    
    # Define los estilos para los encabezados de las columnas.
    header_font = Font(name = 'Arial', size = 12, bold = True, color = "FFFFFF")                    # Estilo/Formato de los encabezados.
    header_fill = PatternFill(start_color = "2C3E50", end_color = "2C3E50", fill_type = "solid")    # Relleno de los encabezados.
    header_alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)      # Alineación de los encabezados.
    
    # Define el estilo de borde para las celdas.
    border_style = Border(
        left = Side(style = 'thin'),
        right = Side(style = 'thin'),
        top = Side(style = 'thin'),
        bottom = Side(style = 'thin')
    )

    # Agrega el título principal (fila 1).
    ws.merge_cells('A1:I1')                 # Combina las celdas de la A1 a la I1 para el título.
    ws['A1'] = titulo                       # Asigna el título a la celda A1.
    ws['A1'].font = title_font              # Aplica el estilo de fuente al título.
    ws['A1'].alignment = title_alignment    # Aplica la alineación al título.
    ws.row_dimensions[1].height = 30        # Ajusta la altura de la fila del título.

    # Ajusta/Asigna un subtítulo con información de búsqueda o listado (fila 2).
    ws.merge_cells('A2:I2')
    
    # Compreuba/Verifica si hay un término/parámetro de búsqueda para definir el subtítulo. Si lo hay, lo incluye en el subtítulo.
    if termino:
        ws['A2'] = f"Búsqueda realizada: '{termino}'"   # Subtítulo con término/parámetro de búsqueda.
    # Si no hay término/parámetro de búsqueda, indica que es un listado completo.
    else:
        ws['A2'] = "Listado completo de productos"  # Subtítulo para listado completo.

    # Aplica estilo al subtítulo.
    ws['A2'].font = Font(name = 'Arial', size = 12)                             # Estilo/Formato del subtítulo.
    ws['A2'].alignment = Alignment(horizontal = "center", vertical = "center")  # Alineación del subtítulo.
    ws.row_dimensions[2].height = 20                                            # Ajusta la altura de la fila del subtítulo.

    # Fecha de generación (fila 3).
    ws.merge_cells('A3:I3')                                                             # Combina las celdas de la A3 a la I3 para la fecha.
    ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"   # Fecha y hora actual.
    ws['A3'].font = Font(name = 'Arial', size = 10, italic = True)                      # Estilo/Formato de la fecha.
    ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")          # Alineación de la fecha.
    ws.row_dimensions[3].height = 18                                                    # Ajusta la altura de la fila de la fecha.

    # Define los encabezados de las columnas (fila 5).
    encabezados = ['ID', 'Nombre', 'Categoría', 'Marca', 'Precio (USD)', 'Descuento (%)', 'Precio sin Desc.', 'Stock', 'Valoración (0-5)']

    # Recorre los encabezados (en la fila 5) para escribirlos en la hoja.
    for col_num, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row = 5, column = col_num)   # Obtiene la celda correspondiente.
        cell.value = encabezado                     # Asigna el valor del encabezado.
        cell.font = header_font                     # Aplica el estilo de fuente a los encabezados.
        cell.fill = header_fill                     # Aplica el relleno a los encabezados.
        cell.alignment = header_alignment           # Aplica la alineación a los encabezados.
        cell.border = border_style                  # Aplica el borde a los encabezados.

    # Ajusta la altura de la fila de los encabezados.
    ws.row_dimensions[5].height = 35

    # Comprueba/Verifica si hay datos de productos para escribir en el Excel. Si no hay datos, escribie el mensaje.
    if not productos:
        ws.merge_cells('A6:I6')                                                             # Combina las celdas de la A6 a la I6 para el mensaje.
        ws['A6'] = "No existen o no se encontraron datos para los criterios especificados." # Mensaje si no hay datos.
        ws['A6'].alignment = Alignment(horizontal = "center", vertical = "center")          # Alineación del mensaje.
        ws['A6'].font = Font(name = 'Arial', size = 11, italic = True)                      # Estilo/Formato del mensaje.
    # Si hay datos de productos, procede a escribirlos en el Excel.
    else:
        # Escribe los datos de productos (desde la fila 6).
        row_num = 6
        # Recorre cada producto en la lista de productos y escribe sus datos en la hoja.
        for producto in productos:
            # Calcula precio sin descuento.
            descuento = producto['descuento']           # Porcentaje de descuento.
            precio_con_descuento = producto['precio']   # Precio con descuento.
            
            # Comprueba/Verifica si el descuento es mayor a 0 para evitar división por cero. Fórmula: precio_sin_descuento = precio / (1 - descuento/100)
            if descuento > 0:
                precio_sin_descuento = precio_con_descuento / (1 - descuento / 100)
            # Si no hay descuento, el precio sin descuento es igual al precio con descuento.
            else:
                precio_sin_descuento = precio_con_descuento

            # Datos a insertar.
            datos_fila = [
                producto['id'],
                producto['nombre'],
                producto['categoria'],
                producto['marca'],
                f"${precio_con_descuento:.2f}",
                f"{descuento}%",
                f"${precio_sin_descuento:.2f}",
                producto['stock'],
                producto['valoracion']
            ]

            # Recorre los datos de la fila para escribirlos en la hoja. Escribe cada dato en su columna correspondiente.
            for col_num, valor in enumerate(datos_fila, 1):
                cell = ws.cell(row = row_num, column = col_num)                         # Obtiene la celda correspondiente.
                cell.value = valor                                                      # Asigna el valor a la celda.
                cell.border = border_style                                              # Aplica el borde a la celda.
                cell.alignment = Alignment(horizontal = "center", vertical = "center")  # Alineación de la celda.
                
                # Resalta las filas alternadas en gris claro.
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color = "ECF0F1", end_color = "ECF0F1", fill_type = "solid")  # Relleno gris claro para filas pares.

            # Incrementa el número de fila para el siguiente producto.            
            row_num += 1

    # Recorre cada columna para ajustar su ancho automáticamente.
    for col_num in range(1, len(encabezados) + 1):
        max_length = 0                              # Inicializa la longitud máxima.
        column_letter = get_column_letter(col_num)  # Obtiene la letra de la columna actual.
        
        # Calcula el ancho basado en el encabezado.
        header_length = len(str(encabezados[col_num - 1]))  # Longitud del encabezado.
        max_length = header_length                          # Inicializa o Iguala la longitud máxima con la del encabezado.
        
        # Recorre las filas para encontrar el valor más largo en la columna. Revisa el ancho de las celdas de datos (desde la fila 6 en adelante).
        for row_num_check in range(6, ws.max_row + 1):
            cell = ws.cell(row = row_num_check, column = col_num)   # Obtiene la celda actual.

            # Intenta obtener la longitud del valor de la celda.
            try:
                # Comprueba/Verifica si la celda tiene valor antes de medir su longitud. En dicho caso, mide la longitud.
                if cell.value:
                    cell_length = len(str(cell.value))  # Longitud del valor de la celda.
                    # Comprueba/Verifica si la longitud de la celda es mayor que la longitud máxima actual. Si es así, actualiza la longitud máxima.
                    if cell_length > max_length:
                        max_length = cell_length    # Actualiza la longitud máxima.
            # En caso de error (por ejemplo, celda vacía), simplemente continúa.
            except:
                pass
        
        # Ajusta el ancho de la columna con un margen adicional.
        adjusted_width = min(max_length + 3, 50)                    # Ancho máximo de 50 para evitar columnas demasiado anchas.
        ws.column_dimensions[column_letter].width = adjusted_width  # Ajusta el ancho de la columna.

    # Congela los paneles (fijar encabezados - desde fila 6).
    ws.freeze_panes = 'A6'

    # Agrega la fila final con total de productos.
    fila_final = ws.max_row + 2                                                             # Fila siguiente a la última con datos.
    ws.merge_cells(f'A{fila_final}:I{fila_final}')                                          # Merge de celdas para el total.
    ws[f'A{fila_final}'] = f"Total de productos incluidos en el informe: {len(productos)}"  # Texto del total.
    ws[f'A{fila_final}'].font = Font(name = 'Arial', size = 10, italic = True, bold = True) # Estilo/Formato de la celda.
    ws[f'A{fila_final}'].alignment = Alignment(horizontal = "center", vertical = "center")  # Alineación de la celda.

    # Guarda el archivo Excel.
    wb.save(ruta_archivo)

    # Mensaje de confirmación por consola.
    print(f"\n\n\tLos datos han sido exportados exitosamente a {ruta_archivo}")

    # Devuelve la ruta del archivo creado (para mostrarla en el main).
    return ruta_archivo